import openpyxl
from work_on_OS import date_picker
from openpyxl.styles import PatternFill
from tkinter import Tk
from tkinter.filedialog import askopenfilename


Tk().withdraw()
file_mame = askopenfilename(title="Выбор файла", filetypes=[("Excel - файлы", ".xlsx")])



def open_file():
    name = file_mame
    wb = openpyxl.load_workbook(name)
    sheet = wb.active
    return sheet


def validation_of_excel_file():
    sheet = open_file()
    for i in range(len(sheet['1'])):
        if "Я подтверждаю, что являюсь совершеннолетним гражданином " \
           "РФ и потребителем никотинсодержащей продукции" in str(sheet['1'][i].value):
            return 0
    return 1


def calculating_the_number_of_items():
    list_of_purchased_products = []
    list_of_presented_products = []
    capsules = "Капсулы"
    capsules_present = "Сменные капсулы Logic Compact"
    count_of_capsules_buy = 0
    count_of_capsules_present = 0
    device = "Logic Compact девайс"
    device_present = "Электронный испаритель Logic Compact"
    count_of_device_buy = 0
    count_of_device_present = 0
    case_present = "Чехол"
    count_of_case_present = 0
    sheet = open_file()
    acc = 1
    while str(sheet['1'][acc].value).find("Подарки") == -1:
        acc = acc + 1
    zero_point = acc - 14
    for i in range(len(sheet['A'])):
        print(f"Проверка {i+1} строки из {len(sheet['A'])}...")
        for j in range(32+zero_point, len(sheet[i + 1])):
            if str(sheet[i + 2][j].value).find(capsules) != -1:
                count_of_capsules_buy += int(sheet[i + 2][j+1].value)
            if str(sheet[i + 2][j].value).find(device) != -1:
                count_of_device_buy += int(sheet[i + 2][j+1].value)
        for j in range(acc, 31+zero_point):
            if str(sheet[i + 2][j].value).find(capsules_present) != -1:
                count_of_capsules_present += int(sheet[i + 2][j+1].value)
            if str(sheet[i + 2][j].value).find(device_present) != -1:
                count_of_device_present += int(sheet[i + 2][j+1].value)
            if str(sheet[i + 2][j].value).find(case_present) != -1:
                count_of_case_present += int(sheet[i + 2][j+1].value)
        one_action_buy = []
        one_action_present = []
        one_action_buy.append(count_of_capsules_buy)
        one_action_buy.append(count_of_device_buy)
        one_action_present.append(count_of_capsules_present)
        one_action_present.append(count_of_device_present)
        one_action_present.append(count_of_case_present)
        count_of_capsules_buy = 0
        count_of_device_buy = 0
        count_of_capsules_present = 0
        count_of_device_present = 0
        count_of_case_present = 0
        list_of_purchased_products.append(one_action_buy)
        list_of_presented_products.append(one_action_present)
    print(list_of_presented_products)
    return list_of_purchased_products, list_of_presented_products, zero_point



def paint_a_cell(cell, color):
    cell.fill = PatternFill(start_color= color,
                            end_color= color,
                            fill_type='solid')


def main_function():
    date = date_picker()
    acc = 0
    four_capsules = 0
    core_device = 0
    one_capsules = 0
    two_capsules = 0
    three_capsules = 0
    case = 0
    name = file_mame
    wb = openpyxl.load_workbook(name)
    sheet = wb.active
    items_and_presents = calculating_the_number_of_items()
    items = items_and_presents[0]
    presents = items_and_presents[1]
    zero_point = items_and_presents[2]
    print("Завершение...")
    for i in range(len(items)-1):
        if items[i][1] == 1:  # Уст обычное + 4 пачке капсул
            if presents[i][0] == 4:
                paint_a_cell(sheet[i + 2][24+zero_point], "0000FF00")
                four_capsules += 1
            else:
                paint_a_cell(sheet[i + 2][24+zero_point], "FFFF0000")
                acc += 1
        elif items[i][0] == 2:  # 2+1
            if presents[i][0] == 1:
                paint_a_cell(sheet[i + 2][24+zero_point], "0000FF00")
                one_capsules += 1
            else:
                paint_a_cell(sheet[i + 2][24+zero_point], "FFFF0000")
                acc += 1
        elif items[i][0] == 3:  # 3 пачке капсул + уст обычное
            if presents[i][1] == 1:
                paint_a_cell(sheet[i + 2][24+zero_point], "0000FF00")
                core_device += 1
            else:
                paint_a_cell(sheet[i + 2][24+zero_point], "FFFF0000")
                acc += 1
        elif items[i][0] == 6:  # 6+3
            if presents[i][0] == 3:
                paint_a_cell(sheet[i + 2][24+zero_point], "0000FF00")
                three_capsules += 1
            else:
                paint_a_cell(sheet[i + 2][24+zero_point], "FFFF0000")
                acc += 1
        elif items[i][0] == 4:  # 4+2
            if presents[i][0] == 2:
                paint_a_cell(sheet[i + 2][24+zero_point], "0000FF00")
                two_capsules += 1
            elif presents[i][2] == 1:
                paint_a_cell(sheet[i + 2][24 + zero_point], "0000FF00")
                case += 1
            else:
                paint_a_cell(sheet[i + 2][24+zero_point], "FFFF0000")
                acc += 1
        elif items[i][0] == 0 and items[i][1] == 0:
            if presents[i][0] == 0 and presents[i][1] == 0:
                pass
            else:
                pass
        else:
            paint_a_cell(sheet[i + 2][24+zero_point], "FFFF0000")
            acc += 1
    for i in range(len(sheet['AG'])-1):
        if str(sheet[i+2][32+zero_point].value).find("NEW Кардхол") != -1:
            paint_a_cell(sheet[i + 2][24+zero_point], "FFFFFF00")

    wb.save("Reports/" + date + '_report.xlsx')
    return four_capsules, core_device, one_capsules, two_capsules, three_capsules, date, acc, case
